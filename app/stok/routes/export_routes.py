# Export Modulu - Excel ve PDF rapor olusturma
import os
import io
from datetime import datetime
from flask import Blueprint, send_file, request
from app import db
from app.stok.models import Urun, StokHareket, Depo
from app.satin_alma.models import SatinAlmaSiparisi, Tedarikci
from app.uretim.models import UretimEmri
from sqlalchemy import func

export_bp = Blueprint('export', __name__)


# ─── EXCEL EXPORTLARI ────────────────────────────────────────────────────────

def _excel_response(wb, dosya_adi):
    """Workbook'u HTTP response olarak gonder"""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=dosya_adi
    )


@export_bp.route('/export/excel/stok')
def excel_stok():
    """Mevcut stok durumu Excel raporu"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return 'openpyxl kurulu degil. pip install openpyxl', 500

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Mevcut Stok'

    # Baslik stili
    baslik_font = Font(bold=True, color='FFFFFF', size=11)
    baslik_fill = PatternFill('solid', fgColor='1a1a2e')
    ince_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    basliklar = ['Stok Kodu', 'Ürün Adı', 'Kategori', 'Birim',
                 'Toplam Giriş', 'Toplam Çıkış', 'Mevcut Stok',
                 'Min Stok', 'Max Stok', 'Durum', 'Alış Fiyatı', 'Satış Fiyatı']

    for col, baslik in enumerate(basliklar, 1):
        cell = ws.cell(row=1, column=col, value=baslik)
        cell.font = baslik_font
        cell.fill = baslik_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = ince_border

    urunler = Urun.query.filter_by(aktif=1).all()
    for row, urun in enumerate(urunler, 2):
        giris = db.session.query(func.sum(StokHareket.miktar)).filter(
            StokHareket.urun_id == urun.id, StokHareket.hareket_tipi == 'giris'
        ).scalar() or 0
        cikis = db.session.query(func.sum(StokHareket.miktar)).filter(
            StokHareket.urun_id == urun.id, StokHareket.hareket_tipi == 'cikis'
        ).scalar() or 0
        mevcut = giris - cikis

        if urun.min_stok > 0 and mevcut <= urun.min_stok:
            durum = 'KRİTİK'
            durum_renk = 'FFCCCC'
        elif urun.max_stok > 0 and mevcut > urun.max_stok:
            durum = 'FAZLA'
            durum_renk = 'CCE5FF'
        else:
            durum = 'NORMAL'
            durum_renk = 'CCFFCC'

        satirveri = [
            urun.stok_kodu, urun.urun_adi,
            urun.kategori.kategori_adi if urun.kategori else '',
            urun.birim, giris, cikis, round(mevcut, 4),
            urun.min_stok, urun.max_stok, durum,
            urun.alis_fiyati, urun.satis_fiyati
        ]
        for col, deger in enumerate(satirveri, 1):
            cell = ws.cell(row=row, column=col, value=deger)
            cell.border = ince_border
            if col == 10:  # Durum kolonu
                cell.fill = PatternFill('solid', fgColor=durum_renk)
            if row % 2 == 0 and col != 10:
                cell.fill = PatternFill('solid', fgColor='F8F9FA')

    # Kolon genislikleri
    genislikler = [15, 35, 20, 10, 15, 15, 15, 12, 12, 10, 12, 12]
    for i, g in enumerate(genislikler, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = g

    tarih = datetime.now().strftime('%Y%m%d_%H%M')
    return _excel_response(wb, f'stok_raporu_{tarih}.xlsx')


@export_bp.route('/export/excel/hareketler')
def excel_hareketler():
    """Stok hareketleri Excel raporu"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return 'openpyxl kurulu degil', 500

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Stok Hareketleri'

    baslik_font = Font(bold=True, color='FFFFFF')
    baslik_fill = PatternFill('solid', fgColor='1a1a2e')

    basliklar = ['Tarih', 'Ürün Kodu', 'Ürün Adı', 'Depo', 'Hareket Tipi', 'Miktar', 'Birim Fiyat', 'Referans', 'Açıklama']
    for col, b in enumerate(basliklar, 1):
        cell = ws.cell(row=1, column=col, value=b)
        cell.font = baslik_font
        cell.fill = baslik_fill

    hareketler = StokHareket.query.order_by(StokHareket.id.desc()).limit(5000).all()
    for row, h in enumerate(hareketler, 2):
        ws.append([
            h.tarih,
            h.urun.stok_kodu if h.urun else '',
            h.urun.urun_adi if h.urun else '',
            h.depo.depo_adi if h.depo else '',
            h.hareket_tipi,
            h.miktar,
            h.birim_fiyat or 0,
            h.referans_tipi or '',
            h.aciklama or ''
        ])
        if h.hareket_tipi == 'giris':
            ws.cell(row=row, column=5).fill = PatternFill('solid', fgColor='CCFFCC')
        elif h.hareket_tipi == 'cikis':
            ws.cell(row=row, column=5).fill = PatternFill('solid', fgColor='FFCCCC')

    for i, g in enumerate([18, 15, 35, 20, 12, 12, 12, 15, 30], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = g

    tarih = datetime.now().strftime('%Y%m%d_%H%M')
    return _excel_response(wb, f'stok_hareketleri_{tarih}.xlsx')


@export_bp.route('/export/excel/siparisler')
def excel_siparisler():
    """Satin alma siparisleri Excel raporu"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return 'openpyxl kurulu degil', 500

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Satın Alma Siparişleri'

    basliklar = ['Sipariş No', 'Tedarikçi', 'Tarih', 'Teslim Tarihi', 'Durum', 'Toplam Tutar', 'Para Birimi']
    ws.append(basliklar)
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1a1a2e')

    siparisler = SatinAlmaSiparisi.query.filter_by(aktif=1).order_by(SatinAlmaSiparisi.id.desc()).all()
    for s in siparisler:
        ws.append([
            s.siparis_no,
            s.tedarikci.unvan if s.tedarikci else '',
            s.siparis_tarihi,
            s.teslim_tarihi or '',
            s.durum,
            s.toplam_tutar,
            s.para_birimi
        ])

    for i, g in enumerate([20, 35, 15, 15, 12, 15, 10], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = g

    tarih = datetime.now().strftime('%Y%m%d_%H%M')
    return _excel_response(wb, f'siparisler_{tarih}.xlsx')


@export_bp.route('/export/excel/uretim')
def excel_uretim():
    """Uretim emirleri Excel raporu"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return 'openpyxl kurulu degil', 500

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Üretim Emirleri'

    basliklar = ['Emir No', 'Ürün', 'Miktar', 'Birim', 'Başlangıç', 'Bitiş', 'Durum', 'Öncelik', 'Operasyon Sayısı']
    ws.append(basliklar)
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1a1a2e')

    emirler = UretimEmri.query.filter_by(aktif=1).order_by(UretimEmri.id.desc()).all()
    for e in emirler:
        ws.append([
            e.emir_no,
            e.urun.urun_adi if e.urun else '',
            e.miktar,
            e.urun.birim if e.urun else '',
            e.planlanan_baslangic or '',
            e.planlanan_bitis or '',
            e.durum,
            e.oncelik,
            len(e.operasyonlar)
        ])

    for i, g in enumerate([18, 35, 10, 8, 15, 15, 12, 10, 12], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = g

    tarih = datetime.now().strftime('%Y%m%d_%H%M')
    return _excel_response(wb, f'uretim_emirleri_{tarih}.xlsx')


# ─── PDF EXPORTLARI ──────────────────────────────────────────────────────────

def _pdf_response(buf, dosya_adi):
    buf.seek(0)
    return send_file(buf, mimetype='application/pdf', as_attachment=True, download_name=dosya_adi)


@export_bp.route('/export/pdf/stok')
def pdf_stok():
    """Kritik stok PDF raporu"""
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError:
        return 'reportlab kurulu degil. pip install reportlab', 500

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                             rightMargin=1*cm, leftMargin=1*cm,
                             topMargin=1.5*cm, bottomMargin=1*cm)

    styles = getSampleStyleSheet()
    baslik_style = ParagraphStyle('baslik', parent=styles['Title'], fontSize=16, spaceAfter=6)
    alt_baslik_style = ParagraphStyle('altbaslik', parent=styles['Normal'], fontSize=10,
                                       textColor=colors.grey, spaceAfter=12)

    elements = []
    elements.append(Paragraph('Mevcut Stok Raporu', baslik_style))
    elements.append(Paragraph(f'Oluşturma Tarihi: {datetime.now().strftime("%d.%m.%Y %H:%M")}', alt_baslik_style))

    # Tablo verisi
    veri = [['Stok Kodu', 'Ürün Adı', 'Birim', 'Toplam Giriş', 'Toplam Çıkış', 'Mevcut', 'Min Stok', 'Durum']]

    urunler = Urun.query.filter_by(aktif=1).all()
    for urun in urunler:
        giris = db.session.query(func.sum(StokHareket.miktar)).filter(
            StokHareket.urun_id == urun.id, StokHareket.hareket_tipi == 'giris'
        ).scalar() or 0
        cikis = db.session.query(func.sum(StokHareket.miktar)).filter(
            StokHareket.urun_id == urun.id, StokHareket.hareket_tipi == 'cikis'
        ).scalar() or 0
        mevcut = giris - cikis
        if urun.min_stok > 0 and mevcut <= urun.min_stok:
            durum = 'KRİTİK'
        elif urun.max_stok > 0 and mevcut > urun.max_stok:
            durum = 'FAZLA'
        else:
            durum = 'NORMAL'
        veri.append([
            urun.stok_kodu, urun.urun_adi[:40], urun.birim,
            f'{giris:.2f}', f'{cikis:.2f}', f'{mevcut:.4f}',
            f'{urun.min_stok:.2f}', durum
        ])

    tablo = Table(veri, colWidths=[3*cm, 8*cm, 2*cm, 3*cm, 3*cm, 3*cm, 2.5*cm, 2.5*cm])
    tablo_stili = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1a2e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (3, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ])
    # Kritik satirlara kirmizi arka plan
    for i, satir in enumerate(veri[1:], 1):
        if satir[-1] == 'KRİTİK':
            tablo_stili.add('BACKGROUND', (7, i), (7, i), colors.HexColor('#FFCCCC'))

    tablo.setStyle(tablo_stili)
    elements.append(tablo)

    doc.build(elements)
    tarih = datetime.now().strftime('%Y%m%d_%H%M')
    return _pdf_response(buf, f'stok_raporu_{tarih}.pdf')


@export_bp.route('/export/pdf/fatura/<int:fatura_id>')
def pdf_fatura(fatura_id):
    """Tek fatura PDF"""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    except ImportError:
        return 'reportlab kurulu degil. pip install reportlab', 500

    from app.fatura.models import Fatura
    fatura = Fatura.query.get_or_404(fatura_id)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                             rightMargin=2*cm, leftMargin=2*cm,
                             topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()

    elements = []
    # Baslik
    tip_txt = 'ALIŞ FATURASI' if fatura.fatura_tipi == 'alis' else 'SATIŞ FATURASI'
    elements.append(Paragraph(tip_txt, ParagraphStyle('bt', fontSize=20, fontName='Helvetica-Bold', spaceAfter=4)))
    elements.append(Paragraph(f'No: {fatura.fatura_no}', ParagraphStyle('fn', fontSize=11, textColor=colors.grey, spaceAfter=2)))
    elements.append(Paragraph(f'Tarih: {fatura.fatura_tarihi}  |  Durum: {fatura.durum.upper()}',
                               ParagraphStyle('fd', fontSize=10, textColor=colors.grey, spaceAfter=16)))
    elements.append(HRFlowable(width='100%', thickness=2, color=colors.HexColor('#1a1a2e')))
    elements.append(Spacer(1, 12))

    # Taraf
    taraf = fatura.tedarikci.unvan if fatura.tedarikci else fatura.musteri_adi or '-'
    elements.append(Paragraph(f'<b>{"TEDARİKÇİ" if fatura.fatura_tipi == "alis" else "MÜŞTERİ"}:</b> {taraf}',
                               ParagraphStyle('tf', fontSize=11, spaceAfter=16)))

    # Kalemler
    veri = [['#', 'Ürün / Açıklama', 'Miktar', 'Birim', 'Birim Fiyat', 'KDV%', 'Toplam']]
    for i, s in enumerate(fatura.satirlar, 1):
        veri.append([
            str(i), s.satir_tanim[:50], f'{s.miktar:.3f}', s.birim,
            f'{s.birim_fiyat:.4f}', f'%{s.kdv_orani:.0f}',
            f'{s.satir_toplam:.2f} {fatura.para_birimi}'
        ])

    tablo = Table(veri, colWidths=[0.8*cm, 7*cm, 2*cm, 1.8*cm, 3*cm, 1.5*cm, 3.5*cm])
    tablo.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1a2e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(tablo)
    elements.append(Spacer(1, 16))

    # Toplam
    toplam_veri = [
        ['Ara Toplam:', f'{fatura.ara_toplam:.2f} {fatura.para_birimi}'],
        ['KDV:', f'{fatura.toplam_kdv:.2f} {fatura.para_birimi}'],
        ['İndirim:', f'-{fatura.toplam_indirim:.2f} {fatura.para_birimi}'],
        ['GENEL TOPLAM:', f'{fatura.genel_toplam:.2f} {fatura.para_birimi}'],
    ]
    toplam_tablo = Table(toplam_veri, colWidths=[4*cm, 4*cm], hAlign='RIGHT')
    toplam_tablo.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ('FONTSIZE', (0, 0), (-1, -2), 9),
        ('FONTSIZE', (0, 3), (-1, 3), 11),
        ('FONTNAME', (0, 3), (-1, 3), 'Helvetica-Bold'),
        ('TOPPADDING', (0, 3), (-1, 3), 8),
        ('LINEABOVE', (0, 3), (-1, 3), 1, colors.black),
    ]))
    elements.append(toplam_tablo)

    doc.build(elements)
    return _pdf_response(buf, f'fatura_{fatura.fatura_no}.pdf')


@export_bp.route('/export/excel/faturalar')
def excel_faturalar():
    """Fatura listesi Excel raporu"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return 'openpyxl kurulu degil', 500

    from app.fatura.models import Fatura

    tip = request.args.get('tip', '')
    tarih_bas = request.args.get('tarih_bas', '')
    tarih_bit = request.args.get('tarih_bit', '')

    q = Fatura.query.filter_by(aktif=1)
    if tip:
        q = q.filter_by(fatura_tipi=tip)
    faturalar = q.order_by(Fatura.id.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Faturalar'

    baslik_font = Font(bold=True, color='FFFFFF', size=10)
    baslik_fill = PatternFill('solid', fgColor='1a1a2e')
    ince_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    basliklar = ['Fatura No', 'Tür', 'Tarih', 'Vade Tarihi', 'Tedarikçi / Müşteri',
                 'Ara Toplam', 'KDV', 'İndirim', 'Genel Toplam', 'Para Birimi', 'Durum', 'Açıklama']

    for col, b in enumerate(basliklar, 1):
        cell = ws.cell(row=1, column=col, value=b)
        cell.font = baslik_font
        cell.fill = baslik_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = ince_border

    durum_renk_map = {
        'taslak': 'E2E8F0', 'kesildi': 'DBEAFE',
        'odendi': 'DCFCE7', 'iptal': 'FEE2E2'
    }

    for row, f in enumerate(faturalar, 2):
        taraf = f.tedarikci.unvan if f.tedarikci else (f.musteri_adi or '')
        veri = [
            f.fatura_no,
            'Alış' if f.fatura_tipi == 'alis' else 'Satış',
            f.fatura_tarihi, f.vade_tarihi or '',
            taraf,
            f.ara_toplam, f.toplam_kdv, f.toplam_indirim, f.genel_toplam,
            f.para_birimi, f.durum, f.aciklama or ''
        ]
        renk = durum_renk_map.get(f.durum, 'FFFFFF')
        for col, deger in enumerate(veri, 1):
            cell = ws.cell(row=row, column=col, value=deger)
            cell.border = ince_border
            if row % 2 == 0:
                cell.fill = PatternFill('solid', fgColor='F8FAFC')
        # Durum kolonu renklendir
        ws.cell(row=row, column=11).fill = PatternFill('solid', fgColor=renk)

    # Toplam satiri
    son_row = len(faturalar) + 2
    ws.cell(row=son_row, column=1, value='TOPLAM').font = Font(bold=True)
    ws.cell(row=son_row, column=9, value=sum(f.genel_toplam for f in faturalar)).font = Font(bold=True)
    ws.cell(row=son_row, column=9).fill = PatternFill('solid', fgColor='FEF9C3')

    genislikler = [20, 8, 14, 14, 35, 14, 12, 12, 16, 10, 12, 30]
    for i, g in enumerate(genislikler, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = g

    tarih = datetime.now().strftime('%Y%m%d_%H%M')
    return _excel_response(wb, f'faturalar_{tarih}.xlsx')


@export_bp.route('/export/excel/mrp')
def excel_mrp():
    """MRP eksik malzeme Excel raporu"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return 'openpyxl kurulu degil', 500

    try:
        from app.uretim.routes.mrp_routes import _mrp_hesapla
        sonuclar, emirler = _mrp_hesapla()
    except Exception as e:
        return f'MRP hesaplama hatasi: {e}', 500

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'MRP Analizi'

    baslik_font = Font(bold=True, color='FFFFFF', size=10)
    baslik_fill = PatternFill('solid', fgColor='1a1a2e')
    ince_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    basliklar = ['Malzeme', 'Stok Kodu', 'Birim', 'İhtiyaç', 'Mevcut Stok',
                 'Eksik Miktar', 'Karşılanma %', 'Varsayılan Tedarikçi', 'Durum']

    for col, b in enumerate(basliklar, 1):
        cell = ws.cell(row=1, column=col, value=b)
        cell.font = baslik_font
        cell.fill = baslik_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = ince_border

    for row, s in enumerate(sonuclar, 2):
        durum = 'EKSİK' if s['eksik'] > 0 else 'YETERLİ'
        veri = [
            s['urun'].urun_adi, s['urun'].stok_kodu, s['urun'].birim,
            s['ihtiyac_toplam'], s['mevcut_stok'], s['eksik'],
            s['saglanabilir_oran'],
            s['tedarikci'].unvan if s['tedarikci'] else 'Tanımsız',
            durum
        ]
        for col, deger in enumerate(veri, 1):
            cell = ws.cell(row=row, column=col, value=deger)
            cell.border = ince_border
        if s['eksik'] > 0:
            ws.cell(row=row, column=6).fill = PatternFill('solid', fgColor='FEE2E2')
            ws.cell(row=row, column=9).fill = PatternFill('solid', fgColor='FEE2E2')
        else:
            ws.cell(row=row, column=9).fill = PatternFill('solid', fgColor='DCFCE7')

    genislikler = [35, 15, 8, 12, 14, 14, 14, 30, 10]
    for i, g in enumerate(genislikler, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = g

    tarih = datetime.now().strftime('%Y%m%d_%H%M')
    return _excel_response(wb, f'mrp_analizi_{tarih}.xlsx')
