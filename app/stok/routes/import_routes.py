"""
Excel'den toplu veri aktarımı — v2
- Türkçe başlıklı şablonlar
- Esnek sütun eşleştirme (Türkçe veya teknik ad)
- Detaylı hata raporu
"""
import os, io
from datetime import datetime
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from app import db

template_klasoru = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'templates')
import_bp = Blueprint('import_modul', __name__, template_folder=template_klasoru)


# ── Sütun Eşleştirme Haritaları ─────────────────────────────────────────────
# Her alan için kabul edilen tüm başlık varyasyonları (küçük harf, boşuksuz)

URUN_KOLON_MAP = {
    'stok_kodu':     ['stok kodu', 'stok_kodu', 'kod', 'ürün kodu', 'urun kodu', 'stok kodu*', 'stok_kodu*'],
    'urun_adi':      ['ürün adı', 'urun adi', 'urun_adi', 'ürün adı*', 'ad', 'adi', 'urun_adi*'],
    'birim':         ['birim', 'ölçü birimi', 'olcu birimi', 'birim*'],
    'kategori_adi':  ['kategori', 'kategori adı', 'kategori_adi', 'grup'],
    'min_stok':      ['minimum stok', 'min stok', 'min_stok', 'minimum', 'kritik stok'],
    'max_stok':      ['maksimum stok', 'max stok', 'max_stok', 'maksimum'],
    'alis_fiyati':   ['alış fiyatı', 'alis fiyati', 'alis_fiyati', 'alış', 'maliyet'],
    'satis_fiyati':  ['satış fiyatı', 'satis fiyati', 'satis_fiyati', 'satış', 'fiyat'],
    'kdv_orani':     ['kdv oranı', 'kdv orani', 'kdv_orani', 'kdv %', 'kdv'],
    'barkod':        ['barkod', 'barkod no', 'ean', 'ean13'],
    'aciklama':      ['açıklama', 'aciklama', 'not', 'notlar', 'detay'],
}

TEDARIKCI_KOLON_MAP = {
    'tedarikci_kodu': ['tedarikçi kodu', 'tedarikci kodu', 'tedarikci_kodu', 'kod', 'tedarikci_kodu*'],
    'unvan':          ['ünvan', 'unvan', 'firma adı', 'firma adi', 'tedarikçi adı', 'unvan*'],
    'vergi_no':       ['vergi no', 'vergi_no', 'vkn', 'vergi numarası'],
    'vergi_dairesi':  ['vergi dairesi', 'vergi_dairesi'],
    'adres':          ['adres', 'adres'],
    'sehir':          ['şehir', 'sehir', 'il'],
    'telefon':        ['telefon', 'tel', 'gsm'],
    'email':          ['e-posta', 'email', 'e-mail', 'mail'],
    'yetkili_kisi':   ['yetkili kişi', 'yetkili kisi', 'yetkili_kisi', 'ilgili kişi', 'yetkili'],
    'odeme_vadesi':   ['ödeme vadesi', 'odeme vadesi', 'odeme_vadesi', 'vade (gün)', 'vade'],
    'para_birimi':    ['para birimi', 'para_birimi', 'döviz'],
}

PERSONEL_KOLON_MAP = {
    'tc_no':           ['tc no', 'tc_no', 'tc kimlik', 'tc', 'kimlik no'],
    'ad':              ['ad', 'adı', 'isim', 'ad*'],
    'soyad':           ['soyad', 'soyadı', 'soyad*'],
    'departman':       ['departman', 'bölüm', 'bolum'],
    'pozisyon':        ['pozisyon', 'görev', 'unvan', 'iş unvanı'],
    'telefon':         ['telefon', 'gsm', 'cep'],
    'email':           ['e-posta', 'email', 'e-mail'],
    'maas':            ['maaş', 'maas', 'ücret', 'ucret', 'brüt maaş'],
    'ise_giris_tarihi':['işe giriş tarihi', 'ise giris', 'ise_giris_tarihi', 'giriş tarihi', 'işe başlama'],
    'dogum_tarihi':    ['doğum tarihi', 'dogum tarihi', 'dogum_tarihi'],
    'cinsiyet':        ['cinsiyet', 'erkek/kadın'],
}


def _kolon_esle(basliklar, kolon_map):
    """
    Excel başlıklarını model alanlarıyla eşleştir.
    Döndürür: {model_alan: excel_indeks}
    """
    eslesme = {}
    for alan, alternatifler in kolon_map.items():
        for i, baslik in enumerate(basliklar):
            # * işareti, fazla boşluk ve özel karakterleri temizle
            temiz = str(baslik).strip().lower()
            temiz = temiz.replace(' *', '').replace('*', '').strip()
            temiz = temiz.replace('  ', ' ')
            if temiz in [a.lower().replace(' *','').replace('*','').strip() for a in alternatifler]:
                eslesme[alan] = i
                break
    return eslesme


def _deger_al(satir_listesi, eslesme, alan, varsayilan=''):
    """Satırdan alan değerini güvenle al"""
    if alan not in eslesme:
        return varsayilan
    idx = eslesme[alan]
    if idx >= len(satir_listesi):
        return varsayilan
    val = satir_listesi[idx]
    if val is None:
        return varsayilan
    return val


def _excel_oku(dosya):
    """Excel oku — başlıklar ve satırlar ayrı döner"""
    import openpyxl
    wb = openpyxl.load_workbook(dosya, data_only=True)
    ws = wb.active
    satirlar = list(ws.values)
    if not satirlar:
        return [], []
    basliklar = [str(b).strip() if b is not None else '' for b in satirlar[0]]
    veriler = []
    for satir in satirlar[1:]:
        satir_listesi = list(satir)
        if all(v is None or str(v).strip() == '' for v in satir_listesi):
            continue
        veriler.append(satir_listesi)
    return basliklar, veriler


def _simdiki_tarih():
    return datetime.now().strftime('%d.%m.%Y')


# ── Şablon İndirme ───────────────────────────────────────────────────────────

@import_bp.route('/import')
def import_panel():
    return render_template('import/import_panel.html')


@import_bp.route('/import/sablon/<tur>')
def sablon_indir(tur):
    """Türkçe başlıklı Excel şablonu indir"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active

    baslik_font  = Font(bold=True, color='FFFFFF', size=11)
    baslik_fill  = PatternFill('solid', fgColor='1e293b')
    zorunlu_fill = PatternFill('solid', fgColor='7f1d1d')   # Koyu kırmızı = zorunlu
    ornek_fill   = PatternFill('solid', fgColor='f0fdf4')
    ince_border  = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    if tur == 'urun':
        ws.title = 'Ürünler'
        # (başlık, zorunlu_mu, örnek_değer, açıklama)
        kolonlar = [
            ('Stok Kodu',       True,  'URN001',          'Benzersiz ürün kodu'),
            ('Ürün Adı',        True,  'Vida M8x20',       'Ürünün tam adı'),
            ('Birim',           True,  'Adet',             'Adet / Kg / Lt / m / m² / m³ / Kutu'),
            ('Kategori',        False, 'Hammadde',         'Yok ise boş bırakın, otomatik oluşur'),
            ('Minimum Stok',    False, '10',               'Kritik stok uyarı seviyesi'),
            ('Maksimum Stok',   False, '100',              'Maksimum stok kapasitesi'),
            ('Alış Fiyatı',     False, '25.50',            'Nokta ile ayrılmış (25.50)'),
            ('Satış Fiyatı',    False, '35.00',            'Nokta ile ayrılmış (35.00)'),
            ('KDV Oranı',       False, '18',               'Sadece rakam (18, 8, 1, 0)'),
            ('Barkod',          False, '8681234567890',    'EAN-13 barkod numarası'),
            ('Açıklama',        False, '',                 'Ürün hakkında notlar'),
        ]
        ornek = ['URN001','Vida M8x20','Adet','Hammadde','10','100','25.50','35.00','18','8681234567890','Test ürünü']
        dosya = 'urun_sablon.xlsx'

    elif tur == 'tedarikci':
        ws.title = 'Tedarikçiler'
        kolonlar = [
            ('Tedarikçi Kodu',  True,  'TDR001',           'Benzersiz tedarikçi kodu'),
            ('Ünvan',           True,  'ABC Ticaret Ltd.',  'Firma ünvanı'),
            ('Vergi No',        False, '1234567890',        '10 haneli VKN'),
            ('Vergi Dairesi',   False, 'Ankara VD',         ''),
            ('Adres',           False, 'Merkez Mah. No:1',  ''),
            ('Şehir',           False, 'Ankara',            ''),
            ('Telefon',         False, '0312 000 0000',     ''),
            ('E-Posta',         False, 'info@abc.com',      ''),
            ('Yetkili Kişi',    False, 'Ali Yılmaz',        ''),
            ('Ödeme Vadesi',    False, '30',                'Gün cinsinden (30, 45, 60)'),
            ('Para Birimi',     False, 'TL',                'TL / USD / EUR'),
        ]
        ornek = ['TDR001','ABC Ticaret Ltd.','1234567890','Ankara VD','Merkez Mah.','Ankara',
                 '0312 000 0000','info@abc.com','Ali Yılmaz','30','TL']
        dosya = 'tedarikci_sablon.xlsx'

    elif tur == 'personel':
        ws.title = 'Personel'
        kolonlar = [
            ('TC Kimlik No',        False, '12345678901',  '11 haneli TC'),
            ('Ad',                  True,  'Ahmet',         ''),
            ('Soyad',               True,  'Yılmaz',        ''),
            ('Departman',           False, 'Üretim',        ''),
            ('Pozisyon',            False, 'Operatör',      ''),
            ('Telefon',             False, '0555 000 0000', ''),
            ('E-Posta',             False, 'ahmet@f.com',   ''),
            ('Maaş',                False, '25000',         'Brüt maaş (TL)'),
            ('İşe Giriş Tarihi',    False, '01.01.2024',    'GG.AA.YYYY formatında'),
            ('Doğum Tarihi',        False, '01.01.1990',    'GG.AA.YYYY formatında'),
            ('Cinsiyet',            False, 'Erkek',         'Erkek veya Kadın'),
        ]
        ornek = ['12345678901','Ahmet','Yılmaz','Üretim','Operatör','0555 000 0000',
                 'ahmet@firma.com','25000','01.01.2024','01.01.1990','Erkek']
        dosya = 'personel_sablon.xlsx'
    else:
        return 'Bilinmeyen şablon', 404

    # ── Başlık satırı yaz
    for col, (baslik, zorunlu, ornek_deger, aciklama) in enumerate(kolonlar, 1):
        cell = ws.cell(row=1, column=col, value=baslik + (' *' if zorunlu else ''))
        cell.font = baslik_font
        cell.fill = zorunlu_fill if zorunlu else baslik_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = ince_border
        ws.row_dimensions[1].height = 22

    # ── Örnek satır yaz
    for col, deger in enumerate(ornek, 1):
        cell = ws.cell(row=2, column=col, value=deger)
        cell.fill = ornek_fill
        cell.border = ince_border

    # ── Sütun genişlikleri
    for col, (baslik, *_) in enumerate(kolonlar, 1):
        harf = ws.cell(row=1, column=col).column_letter
        ws.column_dimensions[harf].width = max(len(baslik) + 6, 16)

    # ── Açıklama sayfası
    ws2 = wb.create_sheet('Açıklamalar')
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 50
    ws2['A1'] = 'SÜTUN'
    ws2['B1'] = 'AÇIKLAMA'
    ws2['A1'].font = Font(bold=True)
    ws2['B1'].font = Font(bold=True)
    ws2.append([''])
    ws2.append(['* işaretli sütunlar', 'Zorunludur, boş bırakılamaz'])
    ws2.append(['Tarih formatı', 'GG.AA.YYYY  (örn: 15.03.2024)'])
    ws2.append(['Fiyat formatı', 'Nokta ile ayrılmış  (örn: 1250.50)'])
    ws2.append([''])
    ws2.append(['ÖNEMLİ', 'İlk satır başlık satırıdır, değiştirmeyin'])
    ws2.append(['', '2. satırdan itibaren veri girin'])
    ws2.append(['', 'Örnek satırı silebilirsiniz'])
    for (ad, _, _, aciklama) in kolonlar:
        if aciklama:
            ws2.append([ad, aciklama])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=dosya)


# ── Veri Yükleme ─────────────────────────────────────────────────────────────

@import_bp.route('/import/yukle/<tur>', methods=['POST'])
def veri_yukle(tur):
    if 'dosya' not in request.files or not request.files['dosya'].filename:
        flash('Dosya seçilmedi', 'danger')
        return redirect(url_for('import_modul.import_panel'))

    dosya = request.files['dosya']
    if not dosya.filename.lower().endswith(('.xlsx', '.xls')):
        flash('Sadece .xlsx veya .xls dosyası kabul edilir', 'danger')
        return redirect(url_for('import_modul.import_panel'))

    try:
        basliklar, veriler = _excel_oku(dosya)
    except Exception as e:
        flash(f'Dosya okunamadı: {e}', 'danger')
        return redirect(url_for('import_modul.import_panel'))

    if not veriler:
        flash('Dosyada veri bulunamadı (2. satırdan itibaren girin)', 'warning')
        return redirect(url_for('import_modul.import_panel'))

    # Sütunları eşleştir
    if tur == 'urun':
        eslesme = _kolon_esle(basliklar, URUN_KOLON_MAP)
        eksik_zorunlu = [k for k in ['stok_kodu','urun_adi','birim'] if k not in eslesme]
    elif tur == 'tedarikci':
        eslesme = _kolon_esle(basliklar, TEDARIKCI_KOLON_MAP)
        eksik_zorunlu = [k for k in ['tedarikci_kodu','unvan'] if k not in eslesme]
    elif tur == 'personel':
        eslesme = _kolon_esle(basliklar, PERSONEL_KOLON_MAP)
        eksik_zorunlu = [k for k in ['ad','soyad'] if k not in eslesme]
    else:
        flash('Bilinmeyen aktarım türü', 'danger')
        return redirect(url_for('import_modul.import_panel'))

    if eksik_zorunlu:
        flash(f'Zorunlu sütunlar bulunamadı: {", ".join(eksik_zorunlu)} — '
              f'Excel başlıklarınızı kontrol edin veya şablonumuzu kullanın', 'danger')
        return redirect(url_for('import_modul.import_panel'))

    basarili, guncellendi, hatali, hatalar = 0, 0, 0, []

    # ── ÜRÜN AKTARIMI
    if tur == 'urun':
        from app.stok.models import Urun
        from app.stok.models.kategori import Kategori
        for i, satir in enumerate(veriler, 2):
            try:
                stok_kodu = str(_deger_al(satir, eslesme, 'stok_kodu', '')).strip()
                urun_adi  = str(_deger_al(satir, eslesme, 'urun_adi', '')).strip()
                birim     = str(_deger_al(satir, eslesme, 'birim', 'Adet')).strip() or 'Adet'

                if not stok_kodu or not urun_adi:
                    hatalar.append(f'Satır {i}: Stok Kodu ve Ürün Adı zorunlu')
                    hatali += 1; continue

                kat_id = None
                kat_adi = str(_deger_al(satir, eslesme, 'kategori_adi', '')).strip()
                if kat_adi:
                    kat = Kategori.query.filter_by(kategori_adi=kat_adi).first()
                    if not kat:
                        kat = Kategori(kategori_adi=kat_adi)
                        db.session.add(kat)
                        db.session.flush()
                    kat_id = kat.id

                def _sayi(alan, varsayilan=0):
                    try: return float(str(_deger_al(satir, eslesme, alan, varsayilan)).replace(',','.') or varsayilan)
                    except: return varsayilan

                mevcut = Urun.query.filter_by(stok_kodu=stok_kodu).first()
                if mevcut:
                    mevcut.urun_adi     = urun_adi
                    mevcut.birim        = birim
                    mevcut.kategori_id  = kat_id or mevcut.kategori_id
                    mevcut.min_stok     = _sayi('min_stok', mevcut.min_stok)
                    mevcut.max_stok     = _sayi('max_stok', mevcut.max_stok)
                    mevcut.alis_fiyati  = _sayi('alis_fiyati', mevcut.alis_fiyati)
                    mevcut.satis_fiyati = _sayi('satis_fiyati', mevcut.satis_fiyati)
                    mevcut.barkod       = str(_deger_al(satir, eslesme, 'barkod', mevcut.barkod or ''))
                    mevcut.aciklama     = str(_deger_al(satir, eslesme, 'aciklama', mevcut.aciklama or ''))
                    guncellendi += 1
                else:
                    u = Urun(
                        stok_kodu   = stok_kodu,
                        urun_adi    = urun_adi,
                        birim       = birim,
                        kategori_id = kat_id,
                        min_stok    = _sayi('min_stok'),
                        max_stok    = _sayi('max_stok'),
                        alis_fiyati = _sayi('alis_fiyati'),
                        satis_fiyati= _sayi('satis_fiyati'),
                        kdv_orani   = _sayi('kdv_orani', 18),
                        barkod      = str(_deger_al(satir, eslesme, 'barkod', '')),
                        aciklama    = str(_deger_al(satir, eslesme, 'aciklama', '')),
                    )
                    db.session.add(u)
                    basarili += 1
            except Exception as e:
                hatalar.append(f'Satır {i}: {str(e)[:100]}')
                hatali += 1

    # ── TEDARİKÇİ AKTARIMI
    elif tur == 'tedarikci':
        from app.satin_alma.models import Tedarikci
        for i, satir in enumerate(veriler, 2):
            try:
                kod   = str(_deger_al(satir, eslesme, 'tedarikci_kodu', '')).strip()
                unvan = str(_deger_al(satir, eslesme, 'unvan', '')).strip()
                if not kod or not unvan:
                    hatalar.append(f'Satır {i}: Tedarikçi Kodu ve Ünvan zorunlu')
                    hatali += 1; continue

                mevcut = Tedarikci.query.filter_by(tedarikci_kodu=kod).first()
                if mevcut:
                    mevcut.unvan = unvan
                    mevcut.telefon = str(_deger_al(satir, eslesme, 'telefon', mevcut.telefon or ''))
                    mevcut.email   = str(_deger_al(satir, eslesme, 'email', mevcut.email or ''))
                    guncellendi += 1
                else:
                    t = Tedarikci(
                        tedarikci_kodu  = kod,
                        unvan           = unvan,
                        vergi_no        = str(_deger_al(satir, eslesme, 'vergi_no', '')),
                        vergi_dairesi   = str(_deger_al(satir, eslesme, 'vergi_dairesi', '')),
                        adres           = str(_deger_al(satir, eslesme, 'adres', '')),
                        sehir           = str(_deger_al(satir, eslesme, 'sehir', '')),
                        telefon         = str(_deger_al(satir, eslesme, 'telefon', '')),
                        email           = str(_deger_al(satir, eslesme, 'email', '')),
                        yetkili_kisi    = str(_deger_al(satir, eslesme, 'yetkili_kisi', '')),
                        odeme_vadesi    = int(float(_deger_al(satir, eslesme, 'odeme_vadesi', 30) or 30)),
                        para_birimi     = str(_deger_al(satir, eslesme, 'para_birimi', 'TL')) or 'TL',
                    )
                    db.session.add(t)
                    basarili += 1
            except Exception as e:
                hatalar.append(f'Satır {i}: {str(e)[:100]}')
                hatali += 1

    # ── PERSONEL AKTARIMI
    elif tur == 'personel':
        from app.ik.models.personel import Personel
        for i, satir in enumerate(veriler, 2):
            try:
                ad    = str(_deger_al(satir, eslesme, 'ad', '')).strip()
                soyad = str(_deger_al(satir, eslesme, 'soyad', '')).strip()
                if not ad or not soyad:
                    hatalar.append(f'Satır {i}: Ad ve Soyad zorunlu')
                    hatali += 1; continue
                p = Personel(
                    tc_no            = str(_deger_al(satir, eslesme, 'tc_no', '')),
                    ad               = ad,
                    soyad            = soyad,
                    departman        = str(_deger_al(satir, eslesme, 'departman', '')),
                    pozisyon         = str(_deger_al(satir, eslesme, 'pozisyon', '')),
                    telefon          = str(_deger_al(satir, eslesme, 'telefon', '')),
                    email            = str(_deger_al(satir, eslesme, 'email', '')),
                    maas             = float(str(_deger_al(satir, eslesme, 'maas', 0)).replace(',','.') or 0),
                    ise_giris_tarihi = str(_deger_al(satir, eslesme, 'ise_giris_tarihi', _simdiki_tarih())),
                    dogum_tarihi     = str(_deger_al(satir, eslesme, 'dogum_tarihi', '')),
                    cinsiyet         = str(_deger_al(satir, eslesme, 'cinsiyet', '')),
                )
                db.session.add(p)
                basarili += 1
            except Exception as e:
                hatalar.append(f'Satır {i}: {str(e)[:100]}')
                hatali += 1

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f'Veritabanı hatası: {e}', 'danger')
        return redirect(url_for('import_modul.import_panel'))

    if basarili:   flash(f'✓ {basarili} yeni kayıt eklendi', 'success')
    if guncellendi:flash(f'↺ {guncellendi} mevcut kayıt güncellendi', 'info')
    if hatali:
        ozet = ' | '.join(hatalar[:5])
        if len(hatalar) > 5: ozet += f' ... (+{len(hatalar)-5} satır daha)'
        flash(f'✗ {hatali} satırda hata: {ozet}', 'warning')

    return redirect(url_for('import_modul.import_panel'))
